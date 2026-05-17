import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


st.set_page_config(
    page_title="Lista de Chequeo GLOBAL GAP",
    layout="wide",
    initial_sidebar_state="expanded"
)

EXCEL_BASE = "norma.xlsx"


def normalizar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    mapa = {}
    for col in df.columns:
        c = str(col).strip().lower()
        if c in ["sección", "seccion"]:
            mapa[col] = "seccion"
        elif c in ["principio"]:
            mapa[col] = "principio"
        elif c in ["criterios", "criterio"]:
            mapa[col] = "criterio"
        elif c in ["nivel"]:
            mapa[col] = "nivel"

    df = df.rename(columns=mapa)
    requeridas = ["seccion", "principio", "criterio", "nivel"]
    faltantes = [c for c in requeridas if c not in df.columns]

    if faltantes:
        raise ValueError(
            f"Faltan columnas requeridas: {faltantes}. "
            "El Excel debe tener: Sección, Principio, Criterios y Nivel."
        )

    return df[requeridas].copy()


def limpiar_texto(valor):
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def es_capitulo(codigo: str) -> bool:
    return bool(re.match(r"^FV-GFS\s+\d{2}$", codigo))


def es_dos_niveles(codigo: str) -> bool:
    return bool(re.match(r"^FV-GFS\s+\d{2}\.\d{2}$", codigo))


def es_tres_niveles(codigo: str) -> bool:
    return bool(re.match(r"^FV-GFS\s+\d{2}\.\d{2}\.\d{2}$", codigo))


def prefijo_dos_niveles(codigo: str) -> str:
    partes = codigo.split()
    if len(partes) < 2:
        return ""
    numeros = partes[1].split(".")
    if len(numeros) >= 2:
        return f"FV-GFS {numeros[0]}.{numeros[1]}"
    return ""


def prefijo_capitulo(codigo: str) -> str:
    partes = codigo.split()
    if len(partes) < 2:
        return ""
    numeros = partes[1].split(".")
    if len(numeros) >= 1:
        return f"FV-GFS {numeros[0]}"
    return ""


@st.cache_data
def cargar_base() -> pd.DataFrame:
    raw_df = pd.read_excel(EXCEL_BASE)
    df = normalizar_columnas(raw_df)

    codigos = df["seccion"].apply(limpiar_texto).tolist()
    prefijos_con_criterios_hijos = {
        prefijo_dos_niveles(codigo)
        for codigo in codigos
        if es_tres_niveles(codigo)
    }

    titulos_subseccion = {}
    titulos_capitulo = {}

    for _, row in df.iterrows():
        seccion = limpiar_texto(row["seccion"])
        principio = limpiar_texto(row["principio"])

        if es_capitulo(seccion):
            titulos_capitulo[seccion] = principio
        elif es_dos_niveles(seccion) and seccion in prefijos_con_criterios_hijos:
            titulos_subseccion[seccion] = principio

    registros = []
    actual = None
    capitulo_actual = ""
    titulo_capitulo_actual = ""

    for _, row in df.iterrows():
        seccion = limpiar_texto(row["seccion"])
        principio = limpiar_texto(row["principio"])
        criterio = limpiar_texto(row["criterio"])
        nivel = limpiar_texto(row["nivel"])

        if es_capitulo(seccion):
            capitulo_actual = seccion
            titulo_capitulo_actual = principio
            continue

        # Si FV-GFS XX.XX tiene criterios hijos FV-GFS XX.XX.01, es subsección, no criterio auditable.
        if es_dos_niveles(seccion) and seccion in prefijos_con_criterios_hijos:
            continue

        # Criterio con tres niveles, por ejemplo FV-GFS 20.01.01.
        if es_tres_niveles(seccion):
            if actual is not None:
                registros.append(actual)

            sub = prefijo_dos_niveles(seccion)
            cap = prefijo_capitulo(seccion)

            actual = {
                "capitulo": cap,
                "titulo_capitulo": titulos_capitulo.get(cap, titulo_capitulo_actual),
                "subseccion": sub,
                "titulo_subseccion": titulos_subseccion.get(sub, ""),
                "seccion": seccion,
                "principio": principio,
                "criterio": criterio,
                "nivel": nivel,
                "cumplimiento": "",
                "estado": "",
                "metodo de auditoria": "",
                "evidencia": "",
                "comentarios": "",
            }
            continue

        # Criterio con dos niveles, por ejemplo FV-GFS 01.01 o FV-GFS 19.08.
        if es_dos_niveles(seccion):
            if actual is not None:
                registros.append(actual)

            cap = prefijo_capitulo(seccion)

            actual = {
                "capitulo": cap,
                "titulo_capitulo": titulos_capitulo.get(cap, titulo_capitulo_actual),
                "subseccion": cap,
                "titulo_subseccion": titulos_capitulo.get(cap, titulo_capitulo_actual),
                "seccion": seccion,
                "principio": principio,
                "criterio": criterio,
                "nivel": nivel,
                "cumplimiento": "",
                "estado": "",
                "metodo de auditoria": "",
                "evidencia": "",
                "comentarios": "",
            }
            continue

        # Fila de continuación del criterio anterior.
        if actual is not None and seccion == "" and principio == "" and criterio != "":
            actual["criterio"] = (actual["criterio"] + "\n" + criterio).strip()
            if nivel and not actual["nivel"]:
                actual["nivel"] = nivel

    if actual is not None:
        registros.append(actual)

    base = pd.DataFrame(registros)

    columnas = [
        "capitulo", "titulo_capitulo",
        "subseccion", "titulo_subseccion",
        "seccion", "principio", "criterio", "nivel",
        "cumplimiento", "estado",
        "metodo de auditoria", "evidencia", "comentarios",
    ]

    for col in columnas:
        if col not in base.columns:
            base[col] = ""

    return base[columnas]


def calcular_resumen(df: pd.DataFrame):
    total = len(df)
    si = (df["estado"] == "Sí lo cumplo").sum()
    no = (df["estado"] == "No lo cumplo").sum()
    na = (df["estado"] == "No aplica").sum()
    pendientes = total - si - no - na
    avance = 0 if total == 0 else round(((si + no + na) / total) * 100, 1)
    cumplimiento = 0 if total == 0 else round((si / max(total - na, 1)) * 100, 1)
    return total, si, no, na, pendientes, avance, cumplimiento


def exportar_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()

    columnas = [
        "capitulo", "titulo_capitulo",
        "subseccion", "titulo_subseccion",
        "seccion", "principio", "criterio", "nivel",
        "cumplimiento", "estado",
        "metodo de auditoria", "evidencia", "comentarios",
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df[columnas].to_excel(writer, index=False, sheet_name="Lista de Chequeo")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb["Lista de Chequeo"]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    widths = {
        "A": 14, "B": 35, "C": 16, "D": 38,
        "E": 18, "F": 45, "G": 90, "H": 20,
        "I": 18, "J": 18, "K": 35, "L": 35, "M": 45,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)

        estado = str(row[9].value).strip().lower()
        if estado == "sí lo cumplo":
            row[9].fill = PatternFill("solid", fgColor="C6EFCE")
        elif estado == "no lo cumplo":
            row[9].fill = PatternFill("solid", fgColor="FFC7CE")
        elif estado == "no aplica":
            row[9].fill = PatternFill("solid", fgColor="D9D9D9")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 95

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()


st.title("Lista de Chequeo GLOBAL GAP")
st.caption("Plantilla digital para completar criterios de auditoría y exportar resultados.")

if "auditoria_df" not in st.session_state:
    st.session_state.auditoria_df = cargar_base()

df = st.session_state.auditoria_df

with st.sidebar:
    st.header("Filtros")

    capitulos = df[["capitulo", "titulo_capitulo"]].drop_duplicates().sort_values("capitulo")
    opciones_capitulo = ["Todos"] + [
        f"{row.capitulo} — {row.titulo_capitulo}" if row.titulo_capitulo else row.capitulo
        for row in capitulos.itertuples(index=False)
    ]

    capitulo_label = st.selectbox("Capítulo", opciones_capitulo)
    capitulo = "Todos" if capitulo_label == "Todos" else capitulo_label.split(" — ")[0]

    if capitulo == "Todos":
        df_subs = df[["subseccion", "titulo_subseccion"]].drop_duplicates()
    else:
        df_subs = df[df["capitulo"] == capitulo][["subseccion", "titulo_subseccion"]].drop_duplicates()

    df_subs = df_subs.sort_values("subseccion")
    opciones_subseccion = ["Todos"] + [
        f"{row.subseccion} — {row.titulo_subseccion}" if row.titulo_subseccion else row.subseccion
        for row in df_subs.itertuples(index=False)
    ]

    subseccion_label = st.selectbox("Subsección", opciones_subseccion)
    subseccion = "Todos" if subseccion_label == "Todos" else subseccion_label.split(" — ")[0]

    niveles = sorted([n for n in df["nivel"].dropna().unique() if str(n).strip()])
    nivel = st.selectbox("Nivel", ["Todos"] + niveles)

    estado = st.selectbox("Estado", ["Todos", "Pendiente", "Sí lo cumplo", "No lo cumplo", "No aplica"])
    busqueda = st.text_input("Buscar")

    st.markdown("---")
    if st.button("Reiniciar formulario"):
        st.session_state.auditoria_df = cargar_base()
        st.rerun()

filtro = pd.Series(True, index=df.index)

if capitulo != "Todos":
    filtro &= df["capitulo"].eq(capitulo)
if subseccion != "Todos":
    filtro &= df["subseccion"].eq(subseccion)
if nivel != "Todos":
    filtro &= df["nivel"].eq(nivel)
if estado == "Pendiente":
    filtro &= df["estado"].eq("")
elif estado != "Todos":
    filtro &= df["estado"].eq(estado)

if busqueda.strip():
    b = busqueda.strip().lower()
    filtro &= (
        df["capitulo"].str.lower().str.contains(b, na=False)
        | df["titulo_capitulo"].str.lower().str.contains(b, na=False)
        | df["subseccion"].str.lower().str.contains(b, na=False)
        | df["titulo_subseccion"].str.lower().str.contains(b, na=False)
        | df["seccion"].str.lower().str.contains(b, na=False)
        | df["principio"].str.lower().str.contains(b, na=False)
        | df["criterio"].str.lower().str.contains(b, na=False)
    )

df_filtrado = df[filtro]
total, si, no, na, pendientes, avance, cumplimiento = calcular_resumen(df)

m1, m2, m3, m4, m5, m6 = st.columns(6)
m1.metric("Criterios", total)
m2.metric("Sí cumple", si)
m3.metric("No cumple", no)
m4.metric("No aplica", na)
m5.metric("Pendientes", pendientes)
m6.metric("Cumplimiento", f"{cumplimiento}%")

st.progress(avance / 100 if avance else 0)
st.caption(f"Avance de llenado: {avance}%")
st.markdown("### Criterios de auditoría")

if df_filtrado.empty:
    st.warning("No hay criterios que coincidan con los filtros.")
else:
    subseccion_anterior = None
    for idx, row in df_filtrado.iterrows():
        if row["subseccion"] != subseccion_anterior:
            st.markdown("---")
            encabezado = row["subseccion"]
            if row["titulo_subseccion"]:
                encabezado += f" — {row['titulo_subseccion']}"
            st.subheader(encabezado)
            subseccion_anterior = row["subseccion"]

        titulo = f"{row['seccion']} — {row['principio']}"
        with st.expander(titulo, expanded=False):
            st.markdown(f"**Nivel:** {row['nivel']}")
            st.markdown("**Criterio:**")
            st.text_area(
                label="criterio",
                value=row["criterio"],
                height=220,
                disabled=True,
                label_visibility="collapsed",
                key=f"criterio_{idx}"
            )

            c1, c2 = st.columns([1, 2])
            with c1:
                df.at[idx, "cumplimiento"] = st.selectbox(
                    "Cumplimiento",
                    ["", "Mayor", "Menor"],
                    index=["", "Mayor", "Menor"].index(df.at[idx, "cumplimiento"])
                    if df.at[idx, "cumplimiento"] in ["", "Mayor", "Menor"] else 0,
                    key=f"cumplimiento_{idx}"
                )

            with c2:
                df.at[idx, "estado"] = st.radio(
                    "Estado",
                    ["", "Sí lo cumplo", "No lo cumplo", "No aplica"],
                    index=["", "Sí lo cumplo", "No lo cumplo", "No aplica"].index(df.at[idx, "estado"])
                    if df.at[idx, "estado"] in ["", "Sí lo cumplo", "No lo cumplo", "No aplica"] else 0,
                    horizontal=True,
                    key=f"estado_{idx}"
                )

            df.at[idx, "metodo de auditoria"] = st.text_area(
                "Método de auditoría",
                value=df.at[idx, "metodo de auditoria"],
                height=90,
                key=f"metodo_{idx}"
            )
            df.at[idx, "evidencia"] = st.text_area(
                "Evidencia",
                value=df.at[idx, "evidencia"],
                height=90,
                key=f"evidencia_{idx}"
            )
            df.at[idx, "comentarios"] = st.text_area(
                "Comentarios",
                value=df.at[idx, "comentarios"],
                height=90,
                key=f"obs_{idx}"
            )

st.session_state.auditoria_df = df
st.markdown("---")

excel_bytes = exportar_excel(st.session_state.auditoria_df)
fecha = datetime.now().strftime("%Y%m%d_%H%M")

st.download_button(
    label="Descargar lista de chequeo completada en Excel",
    data=excel_bytes,
    file_name=f"lista_chequeo_globalgap_completada_{fecha}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
